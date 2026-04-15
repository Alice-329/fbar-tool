#!/usr/bin/env python3
"""
fill_fbar.py  ─  FBAR Auto-Fill GUI
════════════════════════════════════════════════════════════════════════════════
Reads "Tax Info Summary" Excel and a blank NFFBAR.pdf to generate a filled,
Foxit-editable FBAR PDF for the selected calendar year.

Usage (Python):
    python fill_fbar.py

Usage (compiled .exe):
    fill_fbar.exe

Dependencies:  openpyxl  pypdf  xlrd  (pip install openpyxl pypdf xlrd)
════════════════════════════════════════════════════════════════════════════════
"""

import os, re, sys, struct, zlib, threading, tempfile, hashlib
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from pypdf import PdfReader

# ─────────────────────────────────────────────────────────────────────────────
# 0.  PASSWORD PROTECTION
# ─────────────────────────────────────────────────────────────────────────────
_PWD_HASH = "40a3fa414ce9835ac064fe2c943b0b32f3bf3d522c571ad74711d7642fe3dca9"

def _prompt_password():
    """Show a password dialog before the main app opens. Exit if wrong."""
    root = tk.Tk()
    root.withdraw()

    dlg = tk.Toplevel(root)
    dlg.title("FBAR 工具 — 驗證")
    dlg.geometry("320x140")
    dlg.resizable(False, False)
    dlg.grab_set()
    # Centre on screen
    dlg.update_idletasks()
    x = (dlg.winfo_screenwidth()  - 320) // 2
    y = (dlg.winfo_screenheight() - 140) // 2
    dlg.geometry(f"320x140+{x}+{y}")

    tk.Label(dlg, text="請輸入密碼 / Enter Password:", pady=8).pack()
    pwd_var = tk.StringVar()
    entry = tk.Entry(dlg, textvariable=pwd_var, show="*", width=28)
    entry.pack(pady=4)
    entry.focus_set()

    _ok = [False]

    def _do_ok(*_):
        h = hashlib.sha256(pwd_var.get().encode("utf-8")).hexdigest()
        if h == _PWD_HASH:
            _ok[0] = True
            dlg.destroy()
        else:
            messagebox.showerror("錯誤 / Error", "密碼錯誤！\nWrong password!", parent=dlg)
            pwd_var.set("")
            entry.focus_set()

    def _do_cancel(*_):
        dlg.destroy()

    entry.bind("<Return>", _do_ok)
    dlg.protocol("WM_DELETE_WINDOW", _do_cancel)

    btn = tk.Frame(dlg)
    btn.pack(pady=8)
    tk.Button(btn, text="確認 OK",     command=_do_ok,     width=10).pack(side="left",  padx=6)
    tk.Button(btn, text="取消 Cancel", command=_do_cancel, width=10).pack(side="left",  padx=6)

    root.wait_window(dlg)
    root.destroy()

    if not _ok[0]:
        sys.exit(0)

# ─────────────────────────────────────────────────────────────────────────────
# 1.  ADDRESS PARSER
# ─────────────────────────────────────────────────────────────────────────────
# Ordered longest-first so "TAICHUNG CITY" matches before "TAICHUNG"
TAIWAN_CITIES = [
    "NEW TAIPEI CITY", "TAIPEI CITY", "TAICHUNG CITY", "TAINAN CITY",
    "KAOHSIUNG CITY", "TAOYUAN CITY", "HSINCHU CITY", "KEELUNG CITY",
    "CHIAYI CITY",
    "TAIPEI", "TAICHUNG", "TAINAN", "KAOHSIUNG", "TAOYUAN",
    "HSINCHU", "KEELUNG", "CHIAYI", "MIAOLI", "CHANGHUA",
    "NANTOU", "YUNLIN", "PINGTUNG", "HUALIEN", "TAITUNG", "YILAN",
    "PENGHU", "KINMEN", "MATSU",
]

def parse_tw_address(full_addr):
    """
    Parse a combined Taiwan address string into FBAR components.
    Returns (street, city, state, zip_code, country).
    e.g. "NO 49 LANE 80 BO-AI ST NANTUN DIST TAICHUNG 40844 TAIWAN (R.O.C.)"
      -> ("NO 49 LANE 80 BO-AI ST NANTUN DIST", "TAICHUNG", "", "40844", "TW")
    e.g. "NO 82 LIUZHOU ST WANHUA DIST TAIPEI CITY 108012"
      -> ("NO 82 LIUZHOU ST WANHUA DIST", "TAIPEI CITY", "", "108012", "TW")
    """
    if not full_addr:
        return "", "", "", "", "TW"

    addr = str(full_addr).strip().upper()
    # Strip trailing "TAIWAN (R.O.C.)" / "TAIWAN" / "R.O.C." etc.
    addr = re.sub(r'\s*,?\s*TAIWAN\b.*$', '', addr, flags=re.IGNORECASE).strip()
    addr = re.sub(r'\s*,?\s*\(?R\.?O\.?C\.?\)?\s*$', '', addr, flags=re.IGNORECASE).strip()

    # Extract trailing ZIP code (3-6 consecutive digits)
    zip_code = ""
    zip_m = re.search(r'\b(\d{3,6})\s*$', addr)
    if zip_m:
        zip_code = zip_m.group(1)
        addr = addr[:zip_m.start()].strip()

    # Locate city keyword (longest match first)
    city = ""
    street = addr
    for cname in TAIWAN_CITIES:
        idx = addr.rfind(cname)
        if idx >= 0:
            city = cname
            street = addr[:idx].strip()
            # If there is text between city and the already-extracted ZIP,
            # that remainder might itself be the ZIP (rare, but handle it)
            after = addr[idx + len(cname):].strip()
            if after and not zip_code and re.match(r'^\d{3,6}$', after):
                zip_code = after
            break

    return street.strip(), city.strip(), "", zip_code, "TW"


# ─────────────────────────────────────────────────────────────────────────────
# 2.  EXCEL READER
# ─────────────────────────────────────────────────────────────────────────────
def _ensure_xlsx(path):
    """
    If path is already .xlsx, return it unchanged.
    If path is .xls (old Excel 97-2003), convert to a temporary .xlsx file
    using xlrd and return the temp path.  The caller does NOT need to clean up
    (the file is placed in the system temp folder).
    """
    if path.lower().endswith('.xlsx'):
        return path

    # .xls path — convert via xlrd
    try:
        import xlrd
    except ImportError:
        raise ImportError(
            "The selected file is in old .xls format.\n"
            "Please run  '① 安裝套件.bat'  again to install the xlrd package,\n"
            "or save the file as .xlsx in Excel first."
        )

    wb_xls  = xlrd.open_workbook(path, formatting_info=False)
    wb_xlsx = openpyxl.Workbook()
    wb_xlsx.remove(wb_xlsx.active)          # remove the default blank sheet

    for sheet_name in wb_xls.sheet_names():
        ws_xls  = wb_xls.sheet_by_name(sheet_name)
        ws_xlsx = wb_xlsx.create_sheet(title=sheet_name)
        for row in range(ws_xls.nrows):
            for col in range(ws_xls.ncols):
                cell = ws_xls.cell(row, col)
                # xlrd type codes: 0=empty 1=text 2=number 3=date 4=bool 5=error
                ws_xlsx.cell(row + 1, col + 1, cell.value)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    wb_xlsx.save(tmp.name)
    return tmp.name


def _find_sheet(wb, keyword):
    """Return first worksheet whose name contains keyword (case-insensitive)."""
    kw = keyword.lower()
    for name in wb.sheetnames:
        if kw in name.lower():
            return wb[name]
    return None


# ── Dynamic cell search helpers ───────────────────────────────────────────────

def _find_value_by_label(ws, keywords, search_cols=5, max_rows=60):
    """
    Scan the first max_rows rows and search_cols columns for a cell whose text
    contains any keyword (case-insensitive).  Return the value of the first
    non-empty cell to the right of the matched label cell.

    e.g. if "英文姓名" is in row 4 col B, returns the value in row 4 col C.
    """
    for row in range(1, max_rows + 1):
        for col in range(1, search_cols + 1):
            cell_text = re.sub(r'\s+', ' ',
                               str(ws.cell(row, col).value or "")).strip()
            for kw in keywords:
                if kw.lower() in cell_text.lower():
                    # Return first non-empty cell to the right
                    for offset in range(1, 6):
                        val = ws.cell(row, col + offset).value
                        if val is not None and str(val).strip():
                            return val
    return None


def _find_value_in_range(ws, keywords, row_start, row_end, search_cols=5):
    """
    Like _find_value_by_label but restricted to rows [row_start, row_end].
    Used to extract data from a specific filer's section in Basic Information.
    """
    for row in range(row_start, row_end + 1):
        for col in range(1, search_cols + 1):
            cell_text = re.sub(r'\s+', ' ',
                               str(ws.cell(row, col).value or "")).strip()
            for kw in keywords:
                if kw.lower() in cell_text.lower():
                    for offset in range(1, 6):
                        val = ws.cell(row, col + offset).value
                        if val is not None and str(val).strip():
                            return val
    return None


def _find_year_label_row(ws, max_rows=60):
    """
    Find the row that contains year labels like "2020年", "2024年".
    Returns (row_number, {year_str: col_number, ...}) or (None, {}).
    """
    for row in range(1, max_rows + 1):
        year_cols = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val:
                m = re.match(r'^(\d{4})年$', str(val).strip())
                if m:
                    year_cols[m.group(1)] = col
        if len(year_cols) >= 3:          # at least 3 year columns → real label row
            return row, year_cols
    return None, {}


def _find_usd_col(ws, year_row, year_col, year_str):
    """
    Given the column of a year label (e.g. "2024年"), find the corresponding
    USD value column by looking for "YYYY USD" or "YYYYUSD" in the same row
    or the next row, starting from year_col.
    """
    targets = {f"{year_str} USD", f"{year_str}USD"}
    # Search in year_row and year_row+1 from year_col onwards
    for row in (year_row, year_row + 1):
        for col in range(year_col, min(year_col + 4, ws.max_column + 1)):
            val = str(ws.cell(row, col).value or "").strip()
            if val in targets:
                return col
    # Fallback: USD is typically the column immediately after the year label
    return year_col + 1


def _find_header_col(ws, header_row, keywords, search_range=None):
    """
    Search header_row for a column whose cell text contains any keyword.
    search_range: (start_col, end_col) tuple; defaults to entire row.
    Supports CJK headers with embedded newlines (e.g. "持\\n有\\n人") by
    also comparing whitespace-stripped versions of both text and keyword.
    """
    start = search_range[0] if search_range else 1
    end   = search_range[1] if search_range else ws.max_column
    for col in range(start, end + 1):
        text = re.sub(r'\s+', ' ',
                      str(ws.cell(header_row, col).value or "")).strip().upper()
        text_compact = re.sub(r'\s+', '', text)   # strip all whitespace
        for kw in keywords:
            kw_upper   = kw.upper()
            kw_compact = re.sub(r'\s+', '', kw_upper)
            if kw_upper in text or kw_compact in text_compact:
                return col
    return None


def _find_data_start_row(ws, after_row, eng_name_col, max_scan=10):
    """
    Find the first row after after_row that contains actual English account data
    (a non-empty cell in eng_name_col that contains at least one ASCII letter).
    """
    for row in range(after_row + 1, after_row + max_scan + 1):
        val = str(ws.cell(row, eng_name_col).value or "").strip()
        if val and re.search(r'[A-Za-z]', val):
            return row
    return after_row + 1


def detect_years(excel_path):
    """
    Return sorted list of year strings ("2020", "2021", ...) found in Excel.
    Dynamically locates the year-label row in Bank and Brokerage sheets.
    Supports both .xlsx and legacy .xls files.
    """
    xlsx_path = _ensure_xlsx(excel_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    years = set()

    for sheet_keyword in ("Bank account", "Brokerage account"):
        ws = _find_sheet(wb, sheet_keyword)
        if ws:
            _, year_cols = _find_year_label_row(ws)
            years.update(year_cols.keys())

    return sorted(years)


def detect_filers(excel_path):
    """
    Scan the Basic Information sheet and return a list of filer dicts.
    Each dict has keys: chinese_name, first, last, middle, ssn, dob,
                        addr_raw, section_start, section_end.
    Returns [] if only one filer section (or none) is found — caller falls
    back to the existing single-filer parsing in that case.
    """
    try:
        xlsx_path = _ensure_xlsx(excel_path)
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception:
        return []

    basic_ws = _find_sheet(wb, "Basic Information")
    if not basic_ws:
        return []

    # ── Locate "納稅義務人" section markers ────────────────────────────────────
    # Each marker row has "納稅義務人" somewhere in cols A–C and the Chinese name
    # in the first non-empty cell after the marker text.
    section_rows = []  # [(row_num, chinese_name), ...]
    for row in range(1, min(basic_ws.max_row, 120) + 1):
        for col in range(1, 4):
            val = str(basic_ws.cell(row, col).value or "").strip()
            if "納稅義務人" in val and "被撫養人" not in val:
                # Look for the Chinese name in the remaining cols of this row
                chinese_name = ""
                for c in range(col + 1, col + 5):
                    v = str(basic_ws.cell(row, c).value or "").strip()
                    if v:
                        chinese_name = v
                        break
                if not chinese_name:
                    chinese_name = f"申報人{len(section_rows) + 1}"
                section_rows.append((row, chinese_name))
                break

    if len(section_rows) < 2:
        return []   # Single-filer file — no change needed

    # ── Extract filer data from each section ──────────────────────────────────
    filers = []
    for idx, (start_row, chinese_name) in enumerate(section_rows):
        # Section ends just before the next marker (cap at 60 rows per section)
        end_row = (section_rows[idx + 1][0] - 1
                   if idx + 1 < len(section_rows)
                   else min(basic_ws.max_row, start_row + 60))

        def fv(kws):
            return _find_value_in_range(basic_ws, kws, start_row, end_row,
                                        search_cols=5)

        # Prefer explicit First/Last rows (handles "Frist Name" typo too)
        first_val  = fv(["Frist Name", "First Name"])
        last_val   = fv(["Last Name"])
        middle_val = fv(["Middle Name"])

        if first_val or last_val:
            first  = str(first_val  or "").strip().upper()
            last   = str(last_val   or "").strip().upper()
            middle = str(middle_val or "").strip().upper()
        else:
            # Fall back: full English name in "英文姓名" row
            full_en = str(fv(["英文姓名", "English Name"]) or "").strip().upper()
            parts   = full_en.split()
            last    = parts[-1] if parts else ""
            first   = " ".join(parts[:-1]) if len(parts) >= 2 else full_en
            middle  = ""

        ssn_raw = str(fv(["社會安全碼", "Social Security Number", "Tax ID"]) or "")
        ssn     = re.sub(r'[^0-9]', '', ssn_raw.split()[0] if ssn_raw.strip()
                         else "")

        dob_raw = str(fv(["出生年月日", "DOB", "Date of Birth"]) or "").strip()
        dob = re.sub(r'\s*\(.*?\)', '', dob_raw).strip()   # remove "(mm/dd/yyyy)"
        if re.match(r'^\d{4}-\d{2}-\d{2}', dob):
            p = dob.split('-')
            dob = f"{p[1]}/{p[2][:2]}/{p[0]}"

        addr_raw = str(fv(["住址英文", "Current Address", "英文地址"]) or "").strip()

        filers.append({
            "chinese_name":  chinese_name,
            "first":         first,
            "last":          last,
            "middle":        middle,
            "ssn":           ssn,
            "dob":           dob,
            "addr_raw":      addr_raw,
            "section_start": start_row,
            "section_end":   end_row,
        })

    # If a filer has no address, inherit the main filer's address
    main_addr = filers[0].get("addr_raw", "") if filers else ""
    for f in filers[1:]:
        if not f.get("addr_raw"):
            f["addr_raw"] = main_addr

    return filers


def read_excel_data(excel_path, year, holder_name=None):
    """
    Read Excel and return (filer_info dict, accounts list).
    All cell positions are found dynamically — robust to row/column changes.
    Accounts with None balance for the given year are excluded;
    accounts with $0 balance are included.

    holder_name (optional): Chinese name of the specific filer to read
    (e.g. "王朝雍").  When supplied the function uses that filer's section
    of Basic Information AND filters account rows by the 持有人 column.
    """
    year_str  = str(year)
    xlsx_path = _ensure_xlsx(excel_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── Basic Information ─────────────────────────────────────────────────────
    basic_ws = _find_sheet(wb, "Basic Information")
    if not basic_ws:
        raise ValueError("Cannot find 'Basic Information' sheet in Excel")

    # Determine which row range to read for filer data.
    # If holder_name is given and multiple sections exist, target that section.
    bio_start, bio_end = 1, basic_ws.max_row
    if holder_name:
        filer_dicts = detect_filers(excel_path)
        matched = next(
            (f for f in filer_dicts if f["chinese_name"] == holder_name), None)
        if matched:
            bio_start = matched["section_start"]
            bio_end   = matched["section_end"]

    def find_bio(kws):
        v = _find_value_in_range(basic_ws, kws, bio_start, bio_end, search_cols=5)
        # Fallback: whole sheet (e.g. spouse section has no address row)
        if v is None and bio_start > 1:
            v = _find_value_by_label(basic_ws, kws)
        return v

    # Prefer explicit First/Last Name rows (handles "Frist Name" typo)
    first_val  = find_bio(["Frist Name", "First Name"])
    last_val   = find_bio(["Last Name"])
    middle_val = find_bio(["Middle Name"])

    if first_val or last_val:
        first  = str(first_val  or "").strip().upper()
        last   = str(last_val   or "").strip().upper()
        middle = str(middle_val or "").strip().upper()
    else:
        full_name = str(find_bio(["英文姓名", "English Name"]) or "").strip().upper()
        if not full_name:
            raise ValueError(
                "Cannot find English name in Basic Information sheet.\n"
                "Expected a label containing '英文姓名' or 'English Name'.")
        name_parts = full_name.split()
        last   = name_parts[-1] if name_parts else ""
        first  = " ".join(name_parts[:-1]) if len(name_parts) >= 2 else full_name
        middle = ""

    ssn_raw = str(find_bio(["社會安全碼", "Social Security Number", "Tax ID"]) or "").strip()
    ssn     = re.sub(r'[^0-9]', '', ssn_raw.split()[0] if ssn_raw.strip() else "")

    dob_raw = str(find_bio(["出生年月日", "DOB", "Date of Birth"]) or "").strip()
    dob = re.sub(r'\s*\(.*?\)', '', dob_raw).strip()   # strip "(mm/dd/yyyy)"
    if re.match(r'^\d{4}-\d{2}-\d{2}', dob):
        p = dob.split('-')
        dob = f"{p[1]}/{p[2][:2]}/{p[0]}"

    addr_raw = str(find_bio(["住址英文", "Current Address", "英文地址"]) or "").strip()
    f_street, f_city, f_state, f_zip, f_country = parse_tw_address(addr_raw)

    filer = {
        "first":   first,
        "last":    last,
        "middle":  middle,
        "ssn":     ssn,
        "dob":     dob,
        "address": f_street,
        "city":    f_city,
        "state":   f_state,
        "zip":     f_zip,
        "country": f_country,
    }

    accounts = []

    # ── Helper: read accounts from a sheet ───────────────────────────────────
    def _read_accounts(ws, acct_type,
                       eng_name_keywords, addr_keywords, acct_num_keywords,
                       usd_label_patterns, other_desc=""):
        """
        Generic account reader.
        1. Finds the year-label row dynamically.
        2. Finds the USD column for the selected year.
        3. Finds column positions — eng/acct via header keywords,
           address via data content (English text with digits between eng and acct cols).
        4. If holder_name is given and a 持有人 column exists, only reads rows
           for that holder; null-holder rows inherit the last named holder.
        5. Reads data rows until blank.
        """
        yr_row, year_col_map = _find_year_label_row(ws)
        if yr_row is None or year_str not in year_col_map:
            return   # Sheet has no data for this year

        yr_col  = year_col_map[year_str]
        usd_col = _find_usd_col(ws, yr_row, yr_col, year_str)

        # Header row is the row just before the year-label row
        hdr_row = yr_row - 1 if yr_row > 1 else yr_row

        # English name column
        eng_col  = _find_header_col(ws, hdr_row, eng_name_keywords) or \
                   _find_header_col(ws, yr_row,  eng_name_keywords)
        acct_col = _find_header_col(ws, hdr_row, acct_num_keywords) or \
                   _find_header_col(ws, yr_row,  acct_num_keywords)

        # Holder column (新格式: "持有人" / "持\n有\n人")
        holder_col = (
            _find_header_col(ws, hdr_row, ["持有人", "Holder", "Owner"]) or
            _find_header_col(ws, yr_row,  ["持有人", "Holder", "Owner"])
        )

        # Debug: print column headers found
        print(f"  [{ws.title}] hdr_row={hdr_row} yr_row={yr_row} "
              f"eng_col={eng_col} acct_col={acct_col} "
              f"holder_col={holder_col} usd_col={usd_col}")

        if not eng_col or not acct_col:
            print(f"  [{ws.title}] SKIP: could not locate eng or acct column")
            return

        data_start = _find_data_start_row(ws, yr_row, eng_col)

        # Address column: first column between eng_col and acct_col containing
        # English text with digits (a real address).
        addr_col = None
        for col in range(eng_col + 1, acct_col):
            val = str(ws.cell(data_start, col).value or "").strip()
            if (val and re.search(r'[A-Za-z]', val)
                    and re.search(r'\d', val) and len(val) > 10):
                addr_col = col
                break
        print(f"  [{ws.title}] data_start={data_start} addr_col={addr_col}")

        current_holder = None   # tracks last non-null 持有人 value

        for row in range(data_start, ws.max_row + 1):
            # ── Holder tracking & filtering ───────────────────────────────
            if holder_col:
                hv = ws.cell(row, holder_col).value
                if hv and str(hv).strip():
                    current_holder = str(hv).strip()
                # If a specific holder was requested, skip non-matching rows
                if holder_name and current_holder \
                        and current_holder != holder_name:
                    continue

            eng_name = ws.cell(row, eng_col).value
            acct_num = ws.cell(row, acct_col).value
            if not eng_name or not acct_num:
                continue
            eng_name = str(eng_name).strip().replace('\xa0', ' ')
            if not re.search(r'[A-Za-z]', eng_name):
                continue
            acct_str = str(acct_num).strip()
            if not acct_str or acct_str == 'None':
                continue

            usd_raw = ws.cell(row, usd_col).value
            if usd_raw is None:
                continue
            # Detect "unknown" amount (keeps the account but checks 15a)
            _UNKNOWN_VALS = {"unknown", "unknow", "unk", "不知道", "?", "n/a", "na"}
            is_unknown = str(usd_raw).strip().lower() in _UNKNOWN_VALS
            if is_unknown:
                usd_str  = ""
                unkn_val = "X"    # XFA checkbox on-value = "X"
            else:
                try:
                    usd_str  = str(round(float(usd_raw)))
                    unkn_val = ""
                except (TypeError, ValueError):
                    continue

            full_addr = str(ws.cell(row, addr_col).value or "").strip() \
                        if addr_col else ""
            b_street, b_city, b_state, b_zip, b_country = \
                parse_tw_address(full_addr)

            accounts.append({
                "MaximumAccntValue": usd_str,
                "MaximumAccntUnkn":  unkn_val,
                "AccountType":       acct_type,
                "OtherDesc":         other_desc,
                "FinInstName":       eng_name,
                "AccntNumber":       acct_str,
                "Address":           b_street,
                "City":              b_city,
                "State":             b_state,
                "ZIP":               b_zip,
                "Country":           b_country,
            })

    # ── Bank accounts ─────────────────────────────────────────────────────────
    # Note: use "(英文)" as the discriminating keyword — it appears in the
    # English-name column header "銀行名稱(英文)" but NOT in the Chinese-name
    # column "銀行名稱(中文)", so we correctly land on the right column even
    # when both columns share the "銀行名稱" prefix.
    bank_ws = _find_sheet(wb, "Bank account")
    if bank_ws:
        # "(英文)" matches "銀行名稱(英文)" col but NOT "銀行名稱(中文)" col;
        # plain "英文" is a safe fallback since 英文≠中文 in Chinese characters.
        _read_accounts(
            bank_ws, "A",
            eng_name_keywords = ["(英文)", "英文", "English Name", "Bank Name (EN)"],
            addr_keywords     = ["分行地址", "Branch Address", "地址"],
            acct_num_keywords = ["帳號", "Account Number"],
            usd_label_patterns= [f"{year_str} USD", f"{year_str}USD"],
        )

    # ── Brokerage accounts ────────────────────────────────────────────────────
    brok_ws = _find_sheet(wb, "Brokerage account")
    if brok_ws:
        _read_accounts(
            brok_ws, "B",
            eng_name_keywords = ["(英文)", "英文", "English Name", "Institution (EN)"],
            addr_keywords     = ["分行地址", "Branch Address", "地址"],
            acct_num_keywords = ["帳號", "Account Number"],
            usd_label_patterns= [f"{year_str} USD", f"{year_str}USD"],
        )

    # ── Insurance policies ────────────────────────────────────────────────────
    ins_ws = _find_sheet(wb, "Insurance")
    if ins_ws:
        _read_accounts(
            ins_ws, "Z",                                  # Z = "Other" in FBAR dropdown
            eng_name_keywords = ["(英文)", "英文", "English Name", "Institution (EN)"],
            addr_keywords     = ["地址", "Address"],
            acct_num_keywords = ["保單號碼", "Policy", "帳號"],
            usd_label_patterns= [f"{year_str}USD", f"{year_str} USD"],
            other_desc        = "Insurance",
        )

    # ── Other assets ──────────────────────────────────────────────────────────
    other_ws = _find_sheet(wb, "Other assets")
    if other_ws:
        _read_accounts(
            other_ws, "Z",                                # Z = "Other" in FBAR dropdown
            eng_name_keywords = ["(英文)", "英文", "English Name", "Asset Name"],
            addr_keywords     = ["地址", "Address"],
            acct_num_keywords = ["帳號", "Account Number"],
            usd_label_patterns= [f"{year_str}USD", f"{year_str} USD"],
            other_desc        = "Fund",
        )

    return filer, accounts


# ─────────────────────────────────────────────────────────────────────────────
# 3.  DATASETS XML BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def _xe(val):
    """XML-escape a string value so special characters don't break the XFA XML."""
    return (str(val)
            .replace('&',  '&amp;')
            .replace('<',  '&lt;')
            .replace('>',  '&gt;')
            .replace('"',  '&quot;')
            .replace("'",  '&apos;'))


def build_datasets_xml(filer, accounts, year):
    """Return encoded XFA datasets XML bytes."""
    year_str    = str(year)
    filing_name = _xe(f"{filer['first']} {filer['last']} {year_str}")

    # Item 14a: financial interest in Part II accounts (all directly-owned)
    part2_count  = len(accounts)
    fi_25        = "A" if part2_count >= 25 else "B"
    fi_count_xml = str(part2_count) if part2_count >= 25 else ""

    # Item 14b: signature-authority Part IV accounts (individuals: none)
    sig_25 = "B"

    def acct_xml(a):
        ns = 'xmlns:ucc="http://www.fincen.gov/bsa/ucommon-components/2011-06-01"'
        return (
            f'<ffbar:FinAcctOwnedSeparately\n>'
            f'<ffbar:MaximumAccntValue\n>{_xe(a["MaximumAccntValue"])}</ffbar:MaximumAccntValue\n>'
            f'<ffbar:MaximumAccntUnkn\n>{_xe(a.get("MaximumAccntUnkn",""))}</ffbar:MaximumAccntUnkn\n>'
            f'<ffbar:AccountType\n>{_xe(a["AccountType"])}</ffbar:AccountType\n>'
            f'<ffbar:OtherDesc\n>{_xe(a.get("OtherDesc",""))}</ffbar:OtherDesc\n>'
            f'<ffbar:FinInstName\n>{_xe(a["FinInstName"])}</ffbar:FinInstName\n>'
            f'<ffbar:AccntNumber\n>{_xe(a["AccntNumber"])}</ffbar:AccntNumber\n>'
            f'<ffbar:Address\n>'
            f'<ucc:Address {ns}\n>{_xe(a["Address"])}</ucc:Address\n>'
            f'<ucc:City {ns}\n>{_xe(a["City"])}</ucc:City\n>'
            f'<ucc:State {ns}\n>{_xe(a["State"])}</ucc:State\n>'
            f'<ucc:ZIP {ns}\n>{_xe(a["ZIP"])}</ucc:ZIP\n>'
            f'<ucc:Country {ns}\n>{_xe(a["Country"])}</ucc:Country\n>'
            f'</ffbar:Address\n>'
            f'</ffbar:FinAcctOwnedSeparately\n>'
        )

    xml = (
        '<xfa:datasets xmlns:xfa="http://www.xfa.org/schema/xfa-data/1.0/"\n>'
        '<xfa:data xfa:dataNode="dataGroup"\n>'
        '<ffbar:BSAForm xmlns:ffbar="http://www.fincen.gov/bsa/ffbar/2011-06-01"\n>'
        '<ffbar:EFileSubmissionInformation\n>'
        '<est:FilingName xmlns:est="http://www.fincen.gov/bsa/efile-submission-types/2009-01-01"\n>'
        f'{filing_name}'
        '</est:FilingName\n>'
        '<est:FilingType xmlns:est="http://www.fincen.gov/bsa/efile-submission-types/2009-01-01"\n>'
        'FFBAR</est:FilingType\n>'
        '<est:VersionNumber xmlns:est="http://www.fincen.gov/bsa/efile-submission-types/2009-01-01"\n>'
        '1.0</est:VersionNumber\n>'
        '<est:SubmitUrl xmlns:est="http://www.fincen.gov/bsa/efile-submission-types/2009-01-01"\n>'
        'https://bsaefiling1.fincen.treas.gov/submitReport.do</est:SubmitUrl\n>'
        '</ffbar:EFileSubmissionInformation\n>'
        '<ffbar:FilerInformation\n>'
        f'<ffbar:CalendarYear\n>{year_str}</ffbar:CalendarYear\n>'
        '<ffbar:AmendToPriorReports\n/>'
        '<ffbar:TypeOfFiler\n>A</ffbar:TypeOfFiler\n>'
        '<ffbar:FilerOther\n/>'
        f'<ffbar:TIN\n>{_xe(filer["ssn"])}</ffbar:TIN\n>'
        '<ffbar:TINTYPE\n>B</ffbar:TINTYPE\n>'
        '<ffbar:ForeignId\n>'
        '<ffbar:ForeignIdType\n/>'
        '<ffbar:OtherIDDesc\n/>'
        '<ffbar:IdNumber\n/>'
        '<ffbar:IssueCountry\n/>'
        '</ffbar:ForeignId\n>'
        f'<ffbar:DOB\n>{_xe(filer["dob"])}</ffbar:DOB\n>'
        f'<ffbar:LastNameOrNameOfOrg\n>{_xe(filer["last"])}</ffbar:LastNameOrNameOfOrg\n>'
        f'<ffbar:FirstName\n>{_xe(filer["first"])}</ffbar:FirstName\n>'
        f'<ffbar:MiddleName\n>{_xe(filer["middle"])}</ffbar:MiddleName\n>'
        '<ffbar:Suffix\n/>'
        '<ffbar:Address\n>'
        '<ucc:Address xmlns:ucc="http://www.fincen.gov/bsa/ucommon-components/2011-06-01"\n>'
        f'{_xe(filer["address"])}'
        '</ucc:Address\n>'
        '<ucc:City xmlns:ucc="http://www.fincen.gov/bsa/ucommon-components/2011-06-01"\n>'
        f'{_xe(filer["city"])}'
        '</ucc:City\n>'
        '<ucc:State xmlns:ucc="http://www.fincen.gov/bsa/ucommon-components/2011-06-01"\n>'
        f'{_xe(filer["state"])}'
        '</ucc:State\n>'
        '<ucc:ZIP xmlns:ucc="http://www.fincen.gov/bsa/ucommon-components/2011-06-01"\n>'
        f'{_xe(filer["zip"])}'
        '</ucc:ZIP\n>'
        '<ucc:Country xmlns:ucc="http://www.fincen.gov/bsa/ucommon-components/2011-06-01"\n>'
        f'{_xe(filer["country"])}'
        '</ucc:Country\n>'
        '</ffbar:Address\n>'
        f'<ffbar:FIInterestIn25OrMore\n>{fi_25}</ffbar:FIInterestIn25OrMore\n>'
        f'<ffbar:totalNumFIAccnts\n>{fi_count_xml}</ffbar:totalNumFIAccnts\n>'
        f'<ffbar:SigAuth25OrMore\n>{sig_25}</ffbar:SigAuth25OrMore\n>'
        '<ffbar:totalNumSigAuthAccnts\n/>'
        '<ffbar:FilerSignature\n/>'
        '<ffbar:PaidPreparer\n/>'
        '<ffbar:LateFilingReason\n/>'
        '</ffbar:FilerInformation\n>'
        + "".join(acct_xml(a) for a in accounts)
        + '</ffbar:BSAForm\n></xfa:data\n></xfa:datasets\n>'
    )
    return xml.encode('utf-8')


# ─────────────────────────────────────────────────────────────────────────────
# 4.  PDF GENERATOR  (incremental update on blank NFFBAR.pdf)
# ─────────────────────────────────────────────────────────────────────────────

# ── Template patch strings ───────────────────────────────────────────────────
_DOB_OLD = (
    '>this.mandatoryMessage = "Item 5 -  Individual\'s date of birth is required.";\n\n'
    '</script'
)
_DOB_NEW = (
    '>this.mandatoryMessage = "Item 5 -  Individual\'s date of birth is required.";\n'
    'if (dateOfBirthBindField.rawValue != null) '
    '{ this.rawValue = dateOfBirthBindField.rawValue; }\n'
    '</script'
)

_COUNTRY_INIT_OLD = (
    'StatesAndCountriesJS.initializeCountries(this, false); //2nd parm = blankSameAsUS\n'
    'Common.dropDownInitialize(this);\n'
    'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
    '</script'
)
_COUNTRY_INIT_NEW = (
    'StatesAndCountriesJS.initializeCountries(this, false); //2nd parm = blankSameAsUS\n'
    'Common.dropDownInitialize(this);\n'
    'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
    'StatesAndCountriesJS.setStates(this, State, "Required", true, false);\n'
    '</script'
)

_SET_OPT = 'StatesAndCountriesJS.setStates(this, State, "Optional", true, false);\n'
_SET_REQ = 'StatesAndCountriesJS.setStates(this, State, "Required", true, false);\n'

# ── docReady patch: grey State for all non-US accounts ───────────────────
# XFA only greys State via Country's *exit* event (user interaction).
# Pre-filled data never triggers exit; initialize fires BEFORE data binding
# for repeating subforms (Part II–V), so Country.rawValue is null there.
# Solution: patch the *docReady* event, which fires ONCE after ALL data is
# bound, and iterate every Part's Country/State container to apply greying.
#
# Subform structure confirmed from template analysis:
#   Part2:   AddressSub (direct child)
#   Part3:   partSub > AddressSub   AND   PrincipalJointOwner > CitySub
#   Part4/5: partSub > AddressSub   AND   section2 > CitySub
_DOC_READY_JS = (
    'try {\n'
    '  var _B_ = xfa.form.BSAForm;\n'
    '  var _gst_ = function(c) {\n'
    '    if (!c) return;\n'
    '    var co = c.resolveNode("Country");\n'
    '    if (!co) return;\n'
    '    var v = (co.rawValue || "").replace(/ /g, "");\n'
    '    if (!v || v === "US") return;\n'
    '    var s = c.resolveNode("State");\n'
    '    if (!s) return;\n'
    '    s.clearItems(); s.rawValue = null;\n'
    '    if (s.validate) s.validate.nullTest = "disabled";\n'
    '    s.access = "nonInteractive";\n'
    '  };\n'
    '  // Part2: Country/State inside AddressSub (direct child of Part2)\n'
    '  var _p2_ = _B_.resolveNodes("Part2[*]");\n'
    '  if (_p2_) for (var _i_=0; _i_&lt;_p2_.length; _i_++)\n'
    '    _gst_(_p2_.item(_i_).resolveNode("AddressSub"));\n'
    '  // Part3: partSub>AddressSub  +  PrincipalJointOwner>CitySub\n'
    '  var _p3_ = _B_.resolveNodes("Part3[*]");\n'
    '  if (_p3_) for (var _i_=0; _i_&lt;_p3_.length; _i_++) {\n'
    '    var _ps3_ = _p3_.item(_i_).resolveNode("partSub");\n'
    '    if (_ps3_) _gst_(_ps3_.resolveNode("AddressSub"));\n'
    '    var _pj_ = _p3_.item(_i_).resolveNode("PrincipalJointOwner");\n'
    '    if (_pj_) _gst_(_pj_.resolveNode("CitySub"));\n'
    '  }\n'
    '  // Part4/5: partSub>AddressSub  +  section2>CitySub\n'
    '  var _gpt45_ = function(pn) {\n'
    '    if (!pn) return;\n'
    '    for (var _j_=0; _j_&lt;pn.length; _j_++) {\n'
    '      var _pps_ = pn.item(_j_).resolveNode("partSub");\n'
    '      if (_pps_) _gst_(_pps_.resolveNode("AddressSub"));\n'
    '      var _s2_ = pn.item(_j_).resolveNode("section2");\n'
    '      if (_s2_) _gst_(_s2_.resolveNode("CitySub"));\n'
    '    }\n'
    '  };\n'
    '  _gpt45_(_B_.resolveNodes("Part4[*]"));\n'
    '  _gpt45_(_B_.resolveNodes("Part5[*]"));\n'
    '} catch(_e_) {}\n'
)

_DOC_READY_OLD = (
    'activity="docReady" ref="$host" name="event__docReady"\n'
    '><script contentType="application/x-javascript"\n'
    '>Common.docReady();\n'
    '</script\n'
    '></event'
)
_DOC_READY_NEW = (
    'activity="docReady" ref="$host" name="event__docReady"\n'
    '><script contentType="application/x-javascript"\n'
    '>Common.docReady();\n'
    + _DOC_READY_JS
    + '</script\n'
    '></event'
)

_OTHER_PATCHES = [
    ("Part II Country",
     'this.mandatoryMessage = "Part II Item 23 - Country/Region is required.";\n'
     'StatesAndCountriesJS.initializeNoUSCountriesfxcx(this); \n'
     'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
     'Common.dropDownInitialize(this);\n'
     'Common.copyToolTip(Address, this);\n'
     'FFBAR.makePart2Required(secHeaderSub.MaxAcctValue,secHeaderSub.MaxAccntUnknown,secHeaderSub.AccountType,FISub.FIAcctName,AddressSub.AcctNumber,\n'
     '\t\tAddressSub.Address,AddressSub.City,AddressSub.State,AddressSub.ZIP,AddressSub.Country);\n',
     _SET_OPT),

    ("Part III Country",
     'this.mandatoryMessage = "Part III Item 23 - Country/Region is required.";\n'
     'StatesAndCountriesJS.initializeNoUSCountriesfxcx(this);\n'
     'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
     'Common.dropDownInitialize(this);\n'
     'Common.copyToolTip(Address, this);\n'
     'FFBAR.makePartRequired(secHeaderSub.MaxAcctValue,secHeaderSub.MaxAccntUnknown,secHeaderSub.AccountType,FISub.FIAcctName,\n'
     '\t\t\t\t\t\tAddressSub.AcctNumber,AddressSub.Address,AddressSub.City, AddressSub.State,AddressSub.ZIP,AddressSub.Country,AddressSub.NoJointOwner);\n',
     _SET_OPT),

    ("Part IV Country",
     'this.mandatoryMessage = "Part IV Item 23 - Country/Region is required.";\n'
     'StatesAndCountriesJS.initializeNoUSCountriesfxcx(this);\n'
     'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
     'Common.dropDownInitialize(this);\n'
     'Common.copyToolTip(Address, this);\n'
     'FFBAR.makePartRequired(secHeaderSub.MaxAcctValue,secHeaderSub.MaxAccntUnknown,secHeaderSub.AccountType,FISub.FIAcctName,\n'
     '\t\t\t\t\t\tAddressSub.AcctNumber,AddressSub.Address,AddressSub.City, AddressSub.State,AddressSub.ZIP,AddressSub.Country,null);\n',
     _SET_OPT),

    ("Part V Country",
     'this.mandatoryMessage = "Part V Item 23 - Country/Region is required.";\n'
     'StatesAndCountriesJS.initializeNoUSCountriesfxcx(this);\n'
     'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
     'Common.dropDownInitialize(this);\n'
     'Common.copyToolTip(Address, this);\n'
     'FFBAR.makePartRequired(secHeaderSub.MaxAcctValue,secHeaderSub.MaxAccntUnknown,secHeaderSub.AccountType,FISub.FIAcctName,\n'
     '\t\t\t\t\t\tAddressSub.AcctNumber,AddressSub.Address,AddressSub.City, AddressSub.State,AddressSub.ZIP,AddressSub.Country);\n',
     _SET_OPT),

    ("initializeCountries(this,false) Optional",
     'StatesAndCountriesJS.initializeCountries(this,false); //2nd parm = blankSameAsUS\n'
     'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
     'Common.dropDownInitialize(this);\n'
     'Common.copyToolTip(AddSub.Address, this);\n',
     _SET_OPT),

    ("initializeCountries double-newline Optional (x2)",
     'StatesAndCountriesJS.initializeCountries(this, false); //2nd parm = blankSameAsUS\n'
     'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
     'Common.dropDownInitialize(this);\n'
     'Common.copyToolTip(AddSub.Address, this);\n\n',
     _SET_OPT),

    ("Item 59 Paid Preparer Country Required",
     'Item 59 - Country/Region is required.</text\n'
     '></message\n'
     '></validate\n'
     '><event activity="initialize" name="event__initialize"\n'
     '><script contentType="application/x-javascript"\n'
     '>StatesAndCountriesJS.initializeCountries(this, false); //2nd parm = blankSameAsUS\n'
     'StatesAndCountriesJS.setCountryValidatorField(this, CountryValidator);\n'
     'Common.dropDownInitialize(this);\n'
     'Common.copyToolTip(AddSub.Address, this);\n',
     _SET_REQ),
]


def _patch_template(tpl):
    """Apply all template patches and return modified template string."""
    assert _DOB_OLD in tpl, "DOB initialize pattern not found in template"
    tpl = tpl.replace(_DOB_OLD, _DOB_NEW, 1)

    assert _COUNTRY_INIT_OLD in tpl, "Part I Country init pattern not found"
    tpl = tpl.replace(_COUNTRY_INIT_OLD, _COUNTRY_INIT_NEW, 1)

    for desc, old, addition in _OTHER_PATCHES:
        assert old in tpl, f"Pattern not found in template: {desc}"
        tpl = tpl.replace(old, old + addition)

    # ── docReady: grey State for all non-US accounts after full data load ────
    assert _DOC_READY_OLD in tpl, "docReady event pattern not found in template"
    tpl = tpl.replace(_DOC_READY_OLD, _DOC_READY_NEW, 1)

    return tpl


def _stream_obj(obj_id, z_data):
    hdr = (f"{obj_id} 0 obj\n"
           f"<</Filter/FlateDecode/Length {len(z_data)}>>\nstream\n").encode()
    return hdr + z_data + b"\nendstream\nendobj\n"


def generate_fbar_pdf(input_pdf_path, output_pdf_path, filer, accounts, year,
                      log=None):
    """
    Build a filled, Foxit-editable FBAR PDF via PDF incremental update.

    Parameters
    ----------
    input_pdf_path  : path to blank NFFBAR.pdf
    output_pdf_path : destination path for the filled PDF
    filer           : dict returned by read_excel_data()
    accounts        : list of account dicts
    year            : calendar year string or int
    log             : optional callable(str) for progress messages
    """
    def _log(msg):
        if log:
            log(msg)

    _log("Reading blank NFFBAR.pdf ...")
    reader  = PdfReader(input_pdf_path)
    xfa_arr = reader.trailer['/Root']['/AcroForm']['/XFA']
    xfa_map = {}
    for i in range(0, len(xfa_arr), 2):
        xfa_map[xfa_arr[i]] = xfa_arr[i + 1]

    tpl_xml = xfa_map['template'].get_object().get_data().decode('utf-8')
    _log(f"  Template size: {len(tpl_xml):,} chars")

    _log("Patching template (DOB + Country->State graying) ...")
    mod_tpl = _patch_template(tpl_xml)

    _log("Building datasets XML ...")
    datasets = build_datasets_xml(filer, accounts, year)

    new_tpl_z  = zlib.compress(mod_tpl.encode('utf-8'), 9)
    new_data_z = zlib.compress(datasets, 9)
    _log(f"  template compressed: {len(new_tpl_z):,}  datasets: {len(new_data_z):,}")

    # ── Remove /Perms from root object 30 (enables Foxit editing/saving) ─────
    _log("Removing /Perms from PDF root ...")
    with open(input_pdf_path, 'rb') as f:
        orig = f.read()

    obj30_m = re.search(rb'(?<![0-9])30 0 obj\b', orig)
    assert obj30_m, "Object 30 not found in source PDF"
    start30 = obj30_m.start()
    end30_m = re.search(rb'endobj', orig[start30:])
    assert end30_m
    obj30_raw = orig[start30: start30 + end30_m.end()]
    new_root  = re.sub(rb'\s*/Perms\s+49\s+0\s+R', b'', obj30_raw)
    if not new_root.endswith(b'\n'):
        new_root += b'\n'

    # ── Compute offsets ───────────────────────────────────────────────────────
    obj3_b = _stream_obj(3, new_tpl_z)
    obj4_b = _stream_obj(4, new_data_z)

    off3  = len(orig)
    off4  = off3  + len(obj3_b)
    off30 = off4  + len(obj4_b)
    off51 = off30 + len(new_root)

    # ── XRef stream (obj 51): W[1 4 1], Index[3 2 30 1 51 1] ─────────────────
    def xref_n(offset):
        return b'\x01' + struct.pack('>I', offset) + b'\x00'

    entries_raw = xref_n(off3) + xref_n(off4) + xref_n(off30) + xref_n(off51)
    entries_z   = zlib.compress(entries_raw, 9)

    prev_xref = int(re.search(rb'startxref\s+(\d+)', orig).group(1))

    pdf_id = reader.trailer.get('/ID')
    if pdf_id:
        id_str = (f"[<{pdf_id[0].original_bytes.hex()}>"
                  f"<{pdf_id[1].original_bytes.hex()}>]")
    else:
        id_str = ("[<00000000000000000000000000000000>"
                  "<00000000000000000000000000000000>]")

    xref_hdr = (
        f"51 0 obj\n"
        f"<</Type/XRef/Size 52/Prev {prev_xref}"
        f"/Root 30 0 R/Info 28 0 R/ID {id_str}"
        f"/W[1 4 1]/Index[3 2 30 1 51 1]"
        f"/Filter/FlateDecode/Length {len(entries_z)}>>\nstream\n"
    ).encode()

    xref_stream = xref_hdr + entries_z + b"\nendstream\nendobj\n"
    startxref_b = f"startxref\n{off51}\n%%EOF\n".encode()

    # ── Write output ──────────────────────────────────────────────────────────
    _log("Writing output PDF ...")
    final_pdf = orig + obj3_b + obj4_b + new_root + xref_stream + startxref_b
    out_dir = os.path.dirname(output_pdf_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(output_pdf_path, 'wb') as f:
        f.write(final_pdf)
    _log(f"  Saved: {output_pdf_path}  ({len(final_pdf):,} bytes)")


# ─────────────────────────────────────────────────────────────────────────────
# 5.  TKINTER GUI
# ─────────────────────────────────────────────────────────────────────────────
class FbarApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FBAR Auto-Fill Tool")
        self.resizable(True, True)
        self.minsize(640, 540)
        self._filers = []          # list of dicts from detect_filers()
        self._build_ui()
        self._center_window(700, 640)

    # ── Layout ────────────────────────────────────────────────────────────────
    def _build_ui(self):
        PAD = dict(padx=10, pady=4)

        # ─ File paths ─────────────────────────────────────────────────────
        path_frame = ttk.LabelFrame(self, text="  檔案選擇 / File Selection  ")
        path_frame.pack(fill="x", padx=10, pady=(12, 4))
        path_frame.columnconfigure(1, weight=1)

        self._excel_var  = tk.StringVar()
        self._pdf_var    = tk.StringVar()
        self._outdir_var = tk.StringVar()

        file_rows = [
            ("Excel 檔案:",       self._excel_var,  self._browse_excel),
            ("空白 NFFBAR.pdf:",  self._pdf_var,    self._browse_pdf),
            ("輸出資料夾:",        self._outdir_var, self._browse_outdir),
        ]
        for r, (label, var, cmd) in enumerate(file_rows):
            ttk.Label(path_frame, text=label).grid(
                row=r, column=0, sticky="w", padx=8, pady=3)
            ent = ttk.Entry(path_frame, textvariable=var)
            ent.grid(row=r, column=1, sticky="ew", padx=4)
            ttk.Button(path_frame, text="瀏覽...", command=cmd, width=8).grid(
                row=r, column=2, padx=(0, 8))

        # ─ Year selection ─────────────────────────────────────────────────
        yr_frame = ttk.LabelFrame(self, text="  年度選擇 / Year Selection  ")
        yr_frame.pack(fill="x", **PAD)
        self._year_frame_ref = yr_frame   # anchor for person frame insertion

        self._year_var   = tk.StringVar()
        self._year_combo = ttk.Combobox(yr_frame, textvariable=self._year_var,
                                        state="readonly", width=10)
        self._year_combo.pack(side="left", padx=8, pady=6)

        ttk.Button(yr_frame, text="從 Excel 載入年度",
                   command=self._load_years).pack(side="left", padx=4)

        self._year_info_lbl = ttk.Label(yr_frame, text="", foreground="gray")
        self._year_info_lbl.pack(side="left", padx=12)

        # ─ Person selection (hidden until multiple filers detected) ───────
        self._person_frame = ttk.LabelFrame(
            self, text="  申報人選擇 / Filer Selection  ")
        # (not packed yet — shown only when multiple filers found)

        self._person_var   = tk.StringVar()
        self._person_combo = ttk.Combobox(
            self._person_frame, textvariable=self._person_var,
            state="readonly", width=36)
        self._person_combo.pack(side="left", padx=8, pady=6)

        self._person_info_lbl = ttk.Label(
            self._person_frame, text="", foreground="gray")
        self._person_info_lbl.pack(side="left", padx=8)

        self._person_combo.bind("<<ComboboxSelected>>",
                                self._on_person_changed)

        # ─ Preview panel ──────────────────────────────────────────────────
        prev_frame = ttk.LabelFrame(self, text="  資料預覽 / Data Preview  ")
        prev_frame.pack(fill="x", **PAD)
        self._preview_lbl = ttk.Label(
            prev_frame,
            text="（請先選擇 Excel 檔案，然後按「從 Excel 載入年度」）",
            foreground="gray", wraplength=640, justify="left")
        self._preview_lbl.pack(padx=8, pady=5, anchor="w")

        # ─ Generate button ────────────────────────────────────────────────
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=10, pady=2)
        self._gen_btn = ttk.Button(
            btn_frame, text="  產生 FBAR PDF  ",
            command=self._generate)
        self._gen_btn.pack(pady=6, ipadx=10, ipady=4)

        # ─ Log area ───────────────────────────────────────────────────────
        log_frame = ttk.LabelFrame(self, text="  執行記錄 / Log  ")
        log_frame.pack(fill="both", expand=True, padx=10, pady=(2, 10))

        self._log_text = tk.Text(
            log_frame, height=9, state="disabled",
            font=("Consolas", 9), wrap="word",
            bg="#1a1a2e", fg="#e0e0e0", insertbackground="white")
        scrollbar = ttk.Scrollbar(log_frame, command=self._log_text.yview)
        self._log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self._log_text.pack(fill="both", expand=True, padx=2, pady=2)

        # Bind year selection change
        self._year_combo.bind("<<ComboboxSelected>>", self._on_year_changed)

    def _center_window(self, w, h):
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    # ── File browsers ─────────────────────────────────────────────────────────
    def _browse_excel(self):
        p = filedialog.askopenfilename(
            title="選擇 Excel 檔案",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if p:
            self._excel_var.set(p)

    def _browse_pdf(self):
        p = filedialog.askopenfilename(
            title="選擇空白 NFFBAR.pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if p:
            self._pdf_var.set(p)

    def _browse_outdir(self):
        p = filedialog.askdirectory(title="選擇輸出資料夾")
        if p:
            self._outdir_var.set(p)

    # ── Year loading ──────────────────────────────────────────────────────────
    def _load_years(self):
        excel = self._excel_var.get().strip()
        if not excel or not os.path.isfile(excel):
            messagebox.showerror("錯誤", "請先選擇有效的 Excel 檔案")
            return
        try:
            years = detect_years(excel)
        except Exception as e:
            messagebox.showerror("錯誤", f"讀取年度失敗：\n{e}")
            return
        if not years:
            messagebox.showwarning("警告", "未在 Excel 中找到年度資料")
            return
        self._year_combo['values'] = years
        self._year_combo.set(years[-1])
        self._year_info_lbl.config(text=f"共 {len(years)} 個年度")
        self._log(f"已載入年度：{', '.join(years)}  （預設選擇最新年度）")

        # ── Detect multiple filers ──────────────────────────────────────
        try:
            self._filers = detect_filers(excel)
        except Exception:
            self._filers = []

        if self._filers:
            # Build display labels: "王朝雍 (CHAO YUNG WANG)"
            labels = []
            for f in self._filers:
                en = f"{f['first']} {f['last']}".strip()
                label = f"{f['chinese_name']} ({en})" if en else f['chinese_name']
                labels.append(label)
            self._person_combo['values'] = labels
            self._person_combo.set(labels[0])
            n = len(self._filers)
            self._person_info_lbl.config(
                text=f"共 {n} 位申報人，請選擇後再產生 PDF")
            self._log(f"偵測到 {n} 位申報人：{', '.join(f['chinese_name'] for f in self._filers)}")
            # Show person frame (insert between year frame and preview frame)
            self._person_frame.pack(fill="x", padx=10, pady=4,
                                    after=self._year_frame_ref)
        else:
            # Single filer — hide person frame
            self._person_frame.pack_forget()
            self._filers = []
            self._person_var.set("")

        self._refresh_preview()

    def _on_year_changed(self, *_):
        self._refresh_preview()

    def _on_person_changed(self, *_):
        self._refresh_preview()

    # ── Helper: get selected holder_name (None if single-filer) ─────────────
    def _selected_holder(self):
        if not self._filers:
            return None
        label = self._person_var.get().strip()
        if not label:
            return None
        idx = list(self._person_combo['values']).index(label) \
              if label in self._person_combo['values'] else 0
        return self._filers[idx]['chinese_name'] if idx < len(self._filers) \
               else None

    # ── Preview ───────────────────────────────────────────────────────────────
    def _refresh_preview(self):
        excel = self._excel_var.get().strip()
        year  = self._year_var.get().strip()
        if not excel or not year:
            return
        holder = self._selected_holder()
        try:
            filer, accounts = read_excel_data(excel, year, holder_name=holder)
            name = (f"{filer['first']} {filer['middle']} {filer['last']}"
                    .replace('  ', ' ').strip())
            n_bank  = sum(1 for a in accounts if a['AccountType'] == 'A')
            n_brok  = sum(1 for a in accounts if a['AccountType'] == 'B')
            n_other = sum(1 for a in accounts if a['AccountType'] == 'Z')
            lines = [
                f"申報人: {name}    SSN: {filer['ssn']}    DOB: {filer['dob']}",
                f"地址: {filer['address']}, {filer['city']} {filer['zip']} {filer['country']}",
                (f"帳戶數: {len(accounts)} 個  "
                 f"（Bank A: {n_bank}  Brokerage B: {n_brok}  Other Z: {n_other}）"),
            ]
            if len(accounts) >= 25:
                lines.append("! Item 14a: 帳戶數 >= 25，將勾選 Yes")
            self._preview_lbl.config(
                text="\n".join(lines), foreground="#1a1a1a")
        except Exception as e:
            self._preview_lbl.config(
                text=f"預覽失敗：{e}", foreground="red")

    # ── Generate ──────────────────────────────────────────────────────────────
    def _generate(self):
        excel  = self._excel_var.get().strip()
        pdf_in = self._pdf_var.get().strip()
        outdir = self._outdir_var.get().strip()
        year   = self._year_var.get().strip()

        errors = []
        if not excel  or not os.path.isfile(excel):
            errors.append("• 請選擇有效的 Excel 檔案")
        if not pdf_in or not os.path.isfile(pdf_in):
            errors.append("• 請選擇空白 NFFBAR.pdf")
        if not outdir or not os.path.isdir(outdir):
            errors.append("• 請選擇有效的輸出資料夾")
        if not year:
            errors.append("• 請選擇年度（先按「從 Excel 載入年度」）")
        if errors:
            messagebox.showerror("缺少資訊", "\n".join(errors))
            return

        self._gen_btn.config(state="disabled")
        self._log("=" * 52)
        self._log(f"開始產生 {year} 年 FBAR PDF ...")

        holder = self._selected_holder()
        if holder:
            self._log(f"  申報人：{holder}")

        def _worker():
            try:
                filer, accounts = read_excel_data(excel, year,
                                                  holder_name=holder)
                self._log(f"  讀取 Excel 完成：{len(accounts)} 個帳戶")
                for a in accounts:
                    self._log(f"    [{a['AccountType']}] {a['FinInstName']} "
                              f"#{a['AccntNumber']}  USD {a['MaximumAccntValue']}")

                last  = re.sub(r'[^A-Za-z0-9 ]', '', filer['last']).strip()
                first = re.sub(r'[^A-Za-z0-9 ]', '', filer['first']).strip()
                fname = f"{first} {last} FBAR {year}.pdf"
                out_path = os.path.join(outdir, fname)

                generate_fbar_pdf(
                    pdf_in, out_path, filer, accounts, year,
                    log=self._log)

                self.after(0, lambda: self._on_done(out_path))
            except Exception as exc:
                import traceback
                tb = traceback.format_exc()
                self._log(f"ERROR: {exc}")
                self._log(tb)
                self.after(0, lambda: (
                    self._gen_btn.config(state="normal"),
                    messagebox.showerror("產生失敗", str(exc))
                ))

        threading.Thread(target=_worker, daemon=True).start()

    def _on_done(self, out_path):
        self._gen_btn.config(state="normal")
        self._log(f"完成！輸出：{out_path}")
        messagebox.showinfo(
            "完成 / Done",
            f"PDF 已產生：\n\n{out_path}\n\n"
            "請用 Foxit PDF Reader 開啟\n"
            "（支援在 Foxit 中編輯欄位並儲存）。")

    # ── Thread-safe logging ───────────────────────────────────────────────────
    def _log(self, msg):
        def _append():
            self._log_text.config(state="normal")
            self._log_text.insert("end", msg + "\n")
            self._log_text.see("end")
            self._log_text.config(state="disabled")
        self.after(0, _append)


# ─────────────────────────────────────────────────────────────────────────────
# 6.  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    _prompt_password()
    app = FbarApp()
    app.mainloop()
